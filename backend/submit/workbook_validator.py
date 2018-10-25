import pycountry
from django.core.exceptions import ValidationError
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.worksheet import Worksheet

TAB_NAMES = ["Event info", "Clubs", "Fighters"]
TAB_NAMES_TOURNAMENT_TYPE = [
    "Longsword (Steel, Mixed/Men)",
    "Longsword (Steel, Women)",
    "Longsword (Nylon, Mixed)",
    "Rapier &amp; Dagger (Mixed)",
    "Single Rapier (Mixed)",
    "Sabre (Steel, Mixed/Men)",
    "Sword &amp; Buckler (Steel, Mixed)",
    "Sidesword (Steel, Mixed)",
    "Singlestick (Mixed)",
]
COMMON_MISPELLINGS = {
    "lose": "loss",
    "tie": "draw",
}
COUNTRIES_COORDINATES = {
    ("Clubs", "B", ),
    ("Fighters", "C"),
}

# 1. checks if there are core tabs ("Event info", "Clubs", "Fighters") and at least one tournament sheet
# 2. removes empty rows on all sheets
# 3. remove whitespaces (leading, trailing and doubled ones)
# 4. replaces results misspelling ("lose", "tie")
# 5. fixes country names
# 6. sets first tab as active


class WorkbookValidator:
    workbook = None
    sheet_names = []
    errors = []

    def __init__(self, workbook: Workbook) -> None:
        self.workbook = workbook
        self.sheet_names = workbook.sheetnames
        self.errors = []

    def process_workbook(self) -> None:
        """ Main method to process workbook sheets. """

        # todo: what about insensitive cases?
        if not set(TAB_NAMES).issubset(set(self.sheet_names)):
            missing = set(TAB_NAMES) - set(self.sheet_names)
            self.errors.append(ValidationError(f"Missing core tabs {missing}"))

        if not set(TAB_NAMES).intersection(set(self.sheet_names)):
            self.errors.append(ValidationError("Missing at least one tournament sheet"))

        for sheet_title in self.sheet_names:
            self.remove_empty_rows(sheet_title)
            self.parse_cells(sheet_title)
            self.fix_country_name(sheet_title)
            self.verify_results(sheet_title)

        self.verify_users()
        self.set_active_tab()

        if self.errors:
            raise ValidationError(self.errors)

    def remove_empty_rows(self, sheet_title: str) -> None:
        """ Iterate through rows in sheet and remove when all columns have no value. """

        sheet: Worksheet
        sheet = self.workbook[sheet_title]
        max_row = sheet.max_row
        max_col = sheet.max_column
        rows = list(sheet.iter_rows(min_col=1, max_col=max_col, min_row=1, max_row=max_row))
        i = 1
        for row in rows:
            if not any(cell.value for cell in row):
                sheet.delete_rows(i)
            else:
                i += 1

    def parse_cells(self, sheet_title: str) -> None:
        """
        """
        sheet: Worksheet
        sheet = self.workbook[sheet_title]
        max_row = sheet.max_row
        max_col = sheet.max_column
        rows = list(sheet.iter_rows(min_col=1, max_col=max_col, min_row=1, max_row=max_row))
        for row in rows:
            cell: Cell
            for cell in row:
                self.remove_wrong_whitespaces(cell)
                self.fixes_result_name(sheet_title, cell)

    def remove_wrong_whitespaces(self, cell: Cell) -> None:
        """
        Remove trailing and leading whitespaces when cell is string type. It also fixes all duplicated non-standard
        spaces like new line.
        """

        if cell.data_type == "s":
            cell.value = " ".join(cell.value.split())

    def fixes_result_name(self, sheet_title: str, cell: Cell) -> None:
        """
        """

        if sheet_title in TAB_NAMES:
            return

        if cell.column in ['C', 'D']:
            if cell.data_type == "s":
                for wrong, good in COMMON_MISPELLINGS.items():
                    cell.value = cell.value.replace(wrong, good)

    def fix_country_name(self, sheet_title: str) -> None:
        sheet: Worksheet
        sheet = self.workbook[sheet_title]
        for country_coordinate in COUNTRIES_COORDINATES:
            if sheet_title == country_coordinate[0]:
                column = sheet[country_coordinate[1]]
                for cell in column:
                    if cell.value:
                        try:
                            as_name = pycountry.countries.get(name=cell.value.capitalize())
                        except KeyError:
                            as_name = None
                        else:
                            cell.value = as_name.alpha_2

                        try:
                            as_alpha_3 = pycountry.countries.get(alpha_3=cell.value.upper())
                        except KeyError:
                            as_alpha_3 = None
                        else:
                            cell.value = as_alpha_3.alpha_2

                        cell.value = cell.value.upper()

    def verify_users(self) -> None:
        ...

    def verify_results(self, sheet_title: str) -> None:
        if sheet_title in TAB_NAMES:
            return

        sheet: Worksheet
        sheet = self.workbook[sheet_title]

        max_row = sheet.max_row
        rows = list(sheet.iter_rows(min_col=1, max_col=4, min_row=2, max_row=max_row))
        for row in rows:
            f1, f2 = (row[2].value or '').lower(), (row[3].value or '').lower()
            if not ((f1 == 'win' and f2 == 'loss') or (f1 == 'loss' and f2 == 'win') or (f1 == 'draw' and f2 == 'draw')):
                self.errors.append(ValidationError(
                    f"Wrong results at sheet '{sheet_title}' ({f1} and {f2}) - "
                    f"Col {row[0].row}: {row[0].value} vs {row[1].value}"
                ))

    def set_active_tab(self) -> None:
        """
        Set first tab as active by default. workbook.active should be enough but for some reason it does not deactivate
        previously selected one, so we have to iterate through them.
        """

        self.workbook.active = 0
        for sheet in self.workbook:
            if sheet.title == "Event info":
                sheet.sheet_view.tabSelected = True
            else:
                sheet.sheet_view.tabSelected = False
